import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from openpyxl.cell.cell import MergedCell

st.set_page_config(page_title="Amazon 外箱贴自动化工具", layout="wide")

def save_wb(wb):
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

st.title("📦 Amazon 外箱贴自动化工具")
# --- 【新增】功能与使用说明折叠栏 ---
with st.expander("📖 点击查看功能说明与操作指南", expanded=False):
    st.markdown("""
    ### 🌟 功能简介
    本工具专门用于自动化填充亚马逊发货所需的两类表格：
    1. **FBA 发货模板**：自动从计划表中提取 SKU 和数量，解决手动填表的低效与错误。
    2. **装箱信息表**：自动分配分箱数量，并根据“快递”或“海运”模式自动填充外箱尺寸与重量。
    3. **智能换算**：自动识别模板单位（如：磅、英寸），并根据计划表中的数据（KG、CM）进行高精度换算。

    ### 🚀 使用步骤
    * **第一步：生成 FBA 模板**
        1. 分别上传《发货计划表》和亚马逊下载的《原始 SKU 空白模板》。
        2. 点击按钮下载生成好的 FBA 模板，并将其上传至亚马逊后台。
    * **第二步：生成装箱信息表**
        1. 在亚马逊后台下载对应的《装箱信息表》。
        2. 根据物流方式选择 **[快递]** 或 **[海运]**。
        3. 上传下载好的装箱信息表。
        4. 点击按钮下载最终的装箱表，检查无误后上传至亚马逊。

    ### ⚠️ 注意事项
    * **快递模式**：计划表底部必须包含如 `60*50*47` 格式的尺寸行，否则无法提取尺寸。
    * **海运模式**：默认使用标准值（33磅, 24x20x19 in）。
    * **换算标准**：尺寸系数 `0.3937`，重量系数 `2.2046`。
    """)

# --- 文件上传区 ---
# 初始化“记忆”容器，防止上传新表时旧数据丢失
if "plan_data" not in st.session_state:
    st.session_state.plan_data = None

# --- 第一阶段界面 ---
st.subheader("第一步：生成 FBA 发货模板")
c1, c2 = st.columns(2)
with c1:
    plan_file = st.file_uploader("1. 上传《发货计划表》", type=["xlsx"])
with c2:
    fba_template_file = st.file_uploader("2. 上传原始空白《SKU空白模版》", type=["xlsx"])

def safe_write(ws, row, col, value, protect_formula=True):
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if protect_formula and target_cell.value and str(target_cell.value).startswith('='):
                        return
                    target_cell.value = value
                    return
        else:
            if protect_formula and cell.value and str(cell.value).startswith('='):
                return
            cell.value = value
    except:
        pass

def parse_box_range(box_val, qty_val):
    results = []
    try:
        if isinstance(box_val, (int, float)) or (isinstance(box_val, str) and box_val.replace('.','',1).isdigit()):
            results.append((int(float(box_val)), qty_val))
        elif isinstance(box_val, str) and '-' in box_val:
            nums = re.findall(r'\d+', box_val)
            if len(nums) == 2:
                start, end = int(nums[0]), int(nums[1])
                if start <= end:
                    for b in range(start, end + 1):
                        results.append((b, qty_val))
    except: pass
    return results

# --- 逻辑 A：处理 FBA (第一阶段) ---
if plan_file and fba_template_file:
# 1. 先把整张表读进来，别急着删空行
    raw_df = pd.read_excel(plan_file)
    
    # 2. 准备一个名为 box_info 的“小本子”，专门记录尺寸
    box_info = {}   
    
    # 【新增逻辑】：智能雷达！全表扫描寻找真正的“箱规”区域表头
    box_header_row_idx = -1
    box_col_map = {}
    
    for idx, row in raw_df.iterrows():
        row_vals = [str(x).strip() for x in row if pd.notna(x)]
        # 如果这一行同时包含“尺寸”和“箱号”，说明找到了外箱数据区域的表头
        if any("尺寸" in val for val in row_vals) and any("箱号" in val for val in row_vals):
            box_header_row_idx = idx
            for c_idx, cell_val in enumerate(row):
                val_str = str(cell_val).strip()
                if "箱号" in val_str: box_col_map['box_num'] = c_idx
                if "尺寸" in val_str: box_col_map['dim'] = c_idx
                if "重量" in val_str: box_col_map['weight'] = c_idx
            break

    # 模式一：如果找到了外箱专用表头（适用你当前计划表1的格式）
    if box_header_row_idx != -1 and 'dim' in box_col_map:
        for idx in range(box_header_row_idx + 1, len(raw_df)):
            row = raw_df.iloc[idx]
            dim_val = str(row.iloc[box_col_map['dim']]) if pd.notna(row.iloc[box_col_map['dim']]) else ""
            
            if '*' in dim_val:
                dims = [float(d) for d in re.findall(r'\d+\.?\d*', dim_val)]
                if len(dims) == 3:
                    # 提取箱号
                    b_num = None
                    if 'box_num' in box_col_map:
                        b_val = row.iloc[box_col_map['box_num']]
                        if pd.notna(b_val) and str(b_val).replace('.','',1).isdigit():
                            b_num = int(float(b_val))
                    if b_num is None: continue
                    
                    # 提取重量
                    w_val = 0.0
                    if 'weight' in box_col_map:
                        raw_w = row.iloc[box_col_map['weight']]
                        if pd.notna(raw_w) and str(raw_w).replace('.','',1).isdigit():
                            w_val = float(raw_w)
                            
                    box_info[b_num] = {"dim": dims, "weight": w_val}
                    
    else:
        # 模式二：兜底兼容旧格式（以前靠左写的格式）
        for idx, row in raw_df.iterrows():
            col1_val = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            col2_val = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            
            # 【关键防御】：如果第一列有SKU名字，直接跳过，绝不把产品当箱子
            if col1_val and '-' in col1_val and len(col1_val) > 5:
                continue
                
            target_str = col2_val if '*' in col2_val else col1_val
            if '*' in target_str:
                dims = [float(d) for d in re.findall(r'\d+\.?\d*', target_str)]
                if len(dims) == 3:
                    b_num = len(box_info) + 1
                    try:
                        if col1_val.strip().isdigit(): b_num = int(col1_val.strip())
                    except: pass
                    
                    w_val = row.iloc[2] if len(row) > 2 else 0
                    if pd.isna(w_val) and len(row) > 3: w_val = row.iloc[3]
                    w = float(w_val) if pd.notna(w_val) else 0.0
                    box_info[b_num] = {"dim": dims, "weight": w}

    # 把记好的小本子放进 st.session_state
    st.session_state.box_info = box_info
    st.success(f"✅ 成功提取 {len(box_info)} 箱的尺寸和重量信息")

    # 3. 筛选出真正的 SKU 数据
    # 先删掉完全没字的空行
    df = raw_df.dropna(subset=['店铺SKU'])
    # 【重点】把刚才那些带 "*" 号的尺寸行踢出去，不然代码会把它们当成 SKU 去填表，会导致报错
    df = df[~df['店铺SKU'].astype(str).str.contains('\*')] 
    
    # 最后，把这次要发货（数量 > 0）的 SKU 存好
    target_col = [c for c in df.columns if '实际发货数量' in str(c)][0]
    st.session_state.plan_data = df[df[target_col] > 0]

    # ==================== 【强力校验：不通过则拦截下载】 ====================
    used_boxes = set()
    sku_df = st.session_state.plan_data
    box_cols = [c for c in sku_df.columns if '箱号' in str(c)]
    
    for _, row in sku_df.iterrows():
        for col in box_cols:
            if pd.notna(row[col]):
                for b_num, _ in parse_box_range(row[col], 1):
                    used_boxes.add(b_num)
    
    if used_boxes or box_info:
        all_relevant_boxes = used_boxes | set(box_info.keys())
        max_b = max(all_relevant_boxes) if all_relevant_boxes else 0
        expected_seq = set(range(1, max_b + 1))
        
        missing_skus = sorted(list(expected_seq - used_boxes))
        missing_dims = sorted(list(used_boxes - set(box_info.keys())))

        # 情况 A: 箱号断档（如 1, 2, 4）
        if missing_skus:
            st.error(f"❌ **逻辑错误：第 {missing_skus} 箱没有任何产品！**")
            st.warning(f"检测到最大箱号为 {max_b}，但中间箱号分配不连续。系统已拦截文件生成。")
            st.info("💡 请修改《发货计划表》确保箱号连续，然后重新上传。")
            st.stop()  # <--- 关键：强制停止，下方下载按钮不会显示

        if not missing_skus and not missing_dims and max_b > 0:
            st.success(f"✨ 交叉校验通过：1 到 {max_b} 箱配置正确。")
    # =====================================================================

    # 只有上面的 st.stop() 没有被触发，代码才会运行到这里生成文件
    fba_wb = openpyxl.load_workbook(fba_template_file)
    # 1. 提取本次上传表格中实际分配到的箱号
    used_boxes = set()
    sku_df = st.session_state.plan_data
    box_cols = [c for c in sku_df.columns if '箱号' in str(c)]
    
    for _, row in sku_df.iterrows():
        for col in box_cols:
            if pd.notna(row[col]):
                # 调用解析函数处理 "1" 或 "1-3"
                for b_num, _ in parse_box_range(row[col], 1):
                    used_boxes.add(b_num)
    
    # 2. 综合校验：找出最大箱号，检查中间是否有空档
    if used_boxes or box_info:
        # 获取 SKU 分配到的箱号和底部尺寸表定义的箱号的并集
        all_relevant_boxes = used_boxes | set(box_info.keys())
        max_b = max(all_relevant_boxes) if all_relevant_boxes else 0
        expected_seq = set(range(1, max_b + 1))
        
        # 【关键检查】检查 SKU 是否覆盖了 1 到 max_b 的所有箱号
        missing_skus = sorted(list(expected_seq - used_boxes))
        # 检查是否有箱子没填尺寸
        missing_dims = sorted(list(used_boxes - set(box_info.keys())))

        if missing_skus:
            st.error(f"❌ **逻辑错误：第 {missing_skus} 箱没有分配任何产品！**")
            st.warning(f"检测到最大箱号为 {max_b}，但中间的箱号分配不连续。请检查计划表上方区域。")
            # st.stop() # 如果想强制拦截不允许下载，可以取消这一行的注释
            
        if missing_dims:
            st.warning(f"⚠️ **数据缺失：箱号 {missing_dims} 缺少底部的重量尺寸信息！**")

        if not missing_skus and not missing_dims and max_b > 0:
            st.success(f"✨ 校验通过：1 到 {max_b} 箱的 SKU 分配与尺寸信息完整对应。")

    # 2. 立即开始处理 FBA 模板 (在这里创建 fba_wb)
    fba_wb = openpyxl.load_workbook(fba_template_file)
    fba_ws = fba_wb['Template'] if 'Template' in fba_wb.sheetnames else fba_wb.active
    
    # 寻找表头并写入 (主动写入模式)
    header_row_fba, sku_col_fba, qty_col_fba = 0, 1, 2
    for r in range(1, 25):
        row_vals = [str(fba_ws.cell(row=r, column=c).value) for c in range(1, 15)]
        if "Merchant SKU" in row_vals:
            header_row_fba, sku_col_fba = r, row_vals.index("Merchant SKU") + 1
            for idx, val in enumerate(row_vals):
                if "Quantity" in val and "Units" not in val:
                    qty_col_fba = idx + 1
            break

    if header_row_fba > 0:
        if fba_ws.max_row > header_row_fba:
            fba_ws.delete_rows(header_row_fba + 1, fba_ws.max_row)
        curr_row = header_row_fba + 1
        for _, row_data in st.session_state.plan_data.iterrows():
            safe_write(fba_ws, curr_row, sku_col_fba, row_data['店铺SKU'])
            safe_write(fba_ws, curr_row, qty_col_fba, row_data[target_col])
            curr_row += 1
        
        # 3. 数据处理完了，现在显示下载按钮 (此时 fba_wb 已经存在了)
        st.success("✅ FBA 模板处理完成！")
        st.download_button("📥 下载填好的 FBA 模板", save_wb(fba_wb), "FBA_Filled.xlsx")
        full_headers = [
            "Merchant SKU", "Quantity", "Expiration date (MM/DD/YYYY)", 
            "Manufacturing lot code", "Units per box", "Number of boxes", 
            "Box length (in)", "Box width (in)", "Box height (in)", "Box weight (lb)"
        ]

        # 2. 创建一个临时 DataFrame 来存放 10 列数据
        upload_df = pd.DataFrame(columns=full_headers)

        # 3. 把计划表里的 SKU 和 数量 填入前两列
        # 注意：这里直接用你代码里已经定义好的 st.session_state.plan_data 和 target_col
        upload_df["Merchant SKU"] = st.session_state.plan_data['店铺SKU']
        upload_df["Quantity"] = st.session_state.plan_data[target_col]

        # 4. 转换成制表符分隔的 TXT 字符串
        # index=False 表示不需要行号，sep='\t' 是亚马逊要求的制表符格式
        tsv_string = upload_df.to_csv(index=False, sep='\t', encoding='utf-8')

        # 5. 显示第二个下载按钮
        st.download_button(
            label="📄 下载 TXT 格式",
            data=tsv_string,
            file_name="FBA_Upload_Full.txt",
            mime="text/plain"
        )

        txt_df = st.session_state.plan_data[['店铺SKU', target_col]].copy()
        txt_df.columns = ['sku', 'quantity']
        tsv_string = txt_df.to_csv(index=False, sep='\t', encoding='utf-8')


# ==================== 逻辑 B：处理装箱信息表（第二阶段） ====================
if st.session_state.plan_data is not None:
    st.divider() 
    st.subheader("第二步：生成分箱包装信息表")
    
    ship_mode = st.radio("选择配送方式", 
                        ["海运 (默认重量和尺寸)", "快递 (按每箱实际重量和尺寸填写)"], 
                        horizontal=True)
    
    cus_template_file = st.file_uploader("3. 上传从亚马逊下载的《包装箱表》", type=["xlsx"])

    if cus_template_file:
        cus_wb = openpyxl.load_workbook(cus_template_file)
        
        cus_ws = None
        for sheet in cus_wb.worksheets:
            if "包装" in sheet.title:
                cus_ws = sheet
                break
        if not cus_ws:
            cus_ws = cus_wb.worksheets[0]

        # 寻找表头行
        header_row_cus = 0
        for r in range(1, 50):
            row_content = [str(cus_ws.cell(row=r, column=c).value or "").strip().upper() 
                        for c in range(1, 31)]
            if any(k in row_content for k in ["FNSKU", "SKU", "MERCHANT SKU"]):
                header_row_cus = r
                break

        if header_row_cus == 0:
            st.error("❌ 未找到 FNSKU 或 SKU 标题行")
        else:
            col_map = {}
            for c in range(1, cus_ws.max_column + 1):
                cell_value = cus_ws.cell(row=header_row_cus, column=c).value
                col_map[str(cell_value or "").strip()] = c

            sku_col_idx = col_map.get('SKU', col_map.get('Merchant SKU', col_map.get('FNSKU', 1)))
            expected_qty_col_idx = col_map.get('预计数量', 10)

            plan_dict = {}
            for index, row_series in st.session_state.plan_data.iterrows():
                sku_key = str(row_series['店铺SKU']).strip()
                plan_dict[sku_key] = row_series.to_dict()

            target_col = [c for c in st.session_state.plan_data.columns if '实际发货数量' in str(c)][0]

            # 填充预计数量 + 各箱SKU数量
            filled_count = 0
            for curr_row in range(header_row_cus + 1, cus_ws.max_row + 1):
                sku_cell_value = cus_ws.cell(row=curr_row, column=sku_col_idx).value
                if sku_cell_value is None or str(sku_cell_value).strip() in ["", "None"]:
                    break
                    
                sku_in_template = str(sku_cell_value).strip()
                
                if sku_in_template in plan_dict:
                    row_data = plan_dict[sku_in_template]
                    safe_write(cus_ws, curr_row, expected_qty_col_idx, row_data[target_col])
                    
                    # 改进2：无缝提取多组箱号和数量，兼容表格列名后缀（如 箱号.1, 箱号1, 箱号_1 等）
                    box_qty_pairs = []
                    if '箱号' in row_data and '数量' in row_data:
                        box_qty_pairs.append((row_data['箱号'], row_data['数量']))
                    
                    for key in row_data.keys():
                        key_str = str(key)
                        if '箱号' in key_str and key_str != '箱号':
                            suffix = key_str.replace('箱号', '')
                            q_key = f"数量{suffix}"
                            if q_key in row_data:
                                box_qty_pairs.append((row_data[key], row_data[q_key]))
                    
                    for b_val, q_val in box_qty_pairs:
                        if pd.notna(b_val) and pd.notna(q_val):
                            try:
                                if float(q_val) > 0:
                                    for b_num, b_qty in parse_box_range(b_val, q_val):
                                        col_name = f'包装箱 {b_num} 数量'
                                        # 改进3：强制转为标准数字格式，防止后续被当成文本
                                        num_qty = float(b_qty) if float(b_qty) % 1 != 0 else int(float(b_qty))
                                        
                                        if col_name in col_map:
                                            safe_write(cus_ws, curr_row, col_map[col_name], num_qty)
                                            filled_count += 1
                                        else:
                                            for k in col_map:
                                                if f"包装箱 {b_num}" in str(k) and "数量" in str(k):
                                                    safe_write(cus_ws, curr_row, col_map[k], num_qty)
                                                    filled_count += 1
                                                    break
                            except:
                                pass

            # ==================== 【精准正则提取实际箱数】 ====================
            max_box = 0
            for c_name in col_map.keys():
                c_str = str(c_name).strip()
                if re.search(r"包装箱\s*\d+\s*数量", c_str) or re.search(r"Box\s*\d+\s*Quantity", c_str, re.IGNORECASE):
                    match = re.findall(r'\d+', c_str)
                    if match:
                        b_num = int(match[-1])
                        max_box = max(max_box, b_num)
                        
            if max_box == 0:
                max_box = 4 

            # ==================== 【修改：数据完整性校验】 ====================
            empty_box_list = []
            
            for b_num in range(1, max_box + 1):
                current_box_col = None
                col_name_std = f'包装箱 {b_num} 数量'
                
                if col_name_std in col_map:
                    current_box_col = col_map[col_name_std]
                else:
                    for k, v in col_map.items():
                        if f"包装箱 {b_num}" in str(k) and "数量" in str(k):
                            current_box_col = v
                            break
                
                if current_box_col:
                    has_item = False
                    for check_r in range(header_row_cus + 1, cus_ws.max_row + 1):
                        cell_val = cus_ws.cell(row=check_r, column=current_box_col).value
                        # 改进4：打破死板的类型判断，兼容文本格式的数字（例如 "10"）
                        try:
                            if cell_val is not None and str(cell_val).strip() != "" and float(str(cell_val).strip()) > 0:
                                has_item = True
                                break
                        except (ValueError, TypeError):
                            pass
                    
                    if not has_item:
                        empty_box_list.append(b_num)
                else:
                    empty_box_list.append(b_num)

            # ================================================================

            # ==================== 重量和尺寸填充（关键修复：按箱号精确匹配） ====================
            log_rows = {}
            for r in range(header_row_cus + 1, cus_ws.max_row + 1):
                label = str(cus_ws.cell(row=r, column=1).value or "")
                if "重量" in label: log_rows["w"] = (r, label)
                if "宽度" in label: log_rows["wi"] = (r, label)
                if "长度" in label: log_rows["l"] = (r, label)
                if "高度" in label: log_rows["h"] = (r, label)

            # 只填充前 max_box 箱
            actual_filled_boxes = 0  # <--- 新增：准备一个真实填写的计数器

            for c_name, c_idx in list(col_map.items()):
                c_str = str(c_name).strip()
                if not ("包装箱" in c_str or "P1 - B" in c_str):
                    continue

                match = re.findall(r'\d+', c_str)
                if not match:
                    continue

                b_num = int(match[-1])
                if b_num < 1 or b_num > max_box:
                    continue

                # ================= 新增：智能校验该箱子是否分配了产品 =================
                # 1. 获取下方尺寸/重量填写的起始行，划定检查范围，避免把重量值误认为装箱数量
                limit_row = cus_ws.max_row
                if log_rows:
                    limit_row = min([r_idx for r_idx, txt in log_rows.values()])
                
                is_box_used = False
                # 2. 遍历该箱对应的整列，只在“商品数量”区域进行检查
                for check_row in range(header_row_cus + 1, limit_row):
                    cell_val = cus_ws.cell(row=check_row, column=c_idx).value
                    # 如果单元格里有大于 0 的数字，说明这箱确实装了货
                    if isinstance(cell_val, (int, float)) and cell_val > 0:
                        is_box_used = True
                        break
                
                # 3. 如果扫了一圈发现都没装产品，直接跳过当前箱，不填尺寸和重量
                if not is_box_used:
                    continue
                # ====================================================================
                
                actual_filled_boxes += 1

                if ("快递" in ship_mode and 
                    isinstance(st.session_state.box_info, dict) and 
                    b_num in st.session_state.box_info):

                    info = st.session_state.box_info[b_num]
                    w_kg = info.get('weight', 0)
                    l_cm, wi_cm, h_cm = info.get('dim', [0, 0, 0])

                    if "w" in log_rows:
                        r, txt = log_rows["w"]
                        val = w_kg * 2.2046 if any(x in txt for x in ["磅", "lb"]) else w_kg
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "l" in log_rows:
                        r, txt = log_rows["l"]
                        val = l_cm * 0.3937 if any(x in txt for x in ["英寸", "in"]) else l_cm
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "wi" in log_rows:
                        r, txt = log_rows["wi"]
                        val = wi_cm * 0.3937 if any(x in txt for x in ["英寸", "in"]) else wi_cm
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "h" in log_rows:
                        r, txt = log_rows["h"]
                        val = h_cm * 0.3937 if any(x in txt for x in ["英寸", "in"]) else h_cm
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                else:
                    # 海运默认值
                    if "w" in log_rows:
                        r, txt = log_rows["w"]
                        val = 33.0 if any(x in txt for x in ["磅", "lb"]) else 15.0
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "l" in log_rows:
                        r, txt = log_rows["l"]
                        val = 24.0 if any(x in txt for x in ["英寸", "in"]) else 61.0
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "wi" in log_rows:
                        r, txt = log_rows["wi"]
                        val = 20.0 if any(x in txt for x in ["英寸", "in"]) else 51.0
                        safe_write(cus_ws, r, c_idx, round(val, 2))
                    if "h" in log_rows:
                        r, txt = log_rows["h"]
                        val = 19.0 if any(x in txt for x in ["英寸", "in"]) else 48.0
                        safe_write(cus_ws, r, c_idx, round(val, 2))

            st.success(f"✅ 装箱信息表处理完成！已自动过滤空箱，实际填充 {actual_filled_boxes} 箱的重量和尺寸")
            st.download_button("📥 下载填好的装箱信息表", save_wb(cus_wb), "Packing_List_Filled.xlsx")